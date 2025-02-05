#!/usr/bin/python -u
# -*- coding: UTF-8 -*-
import argparse
import os.path
import sys
import time
import yaml
from alive_progress import alive_bar, styles
from yaspin import yaspin
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string, get_column_letter
from progress.bar import Bar
from validator import *
import faulthandler


def is_valid(type, value, coordinate=None, errors=None, value2=None):
    """Takes the validation type assigned to the cell,
    cell value,  coordinates of the cell, errors of previous validation break list
    """
    # Assigning each class to the validation type
    classmap = {
        'NotBlank':    NotBlankValidator.NotBlankValidator,
        'Type':        TypeValidator.TypeValidator,
        'Length':      LengthValidator.LengthValidator,
        'Regex':       RegexValidator.RegexValidator,
        'Email':       EmailValidator.EmailValidator,
        'Choice':      ChoiceValidator.ChoiceValidator,
        'Date':        DateTimeValidator.DateTimeValidator,
        'ExcelDate':   ExcelDateValidator.ExcelDateValidator,
        'Country':     CountryValidator.CountryValidator,
        'Conditional': ConditionalValidator.ConditionalValidator,
        'Order':       OrderValidator.OrderValidator
    }
    violations = []
    # name is the validation type name (NotBlank, Regex, Length ,.... etc)
    # data is the value assigned by the user to be validated ( 3 chars , regex pattern , ... etc)
    name = list(type.keys())[0]
    data = list(type.values())[0]
    validator = classmap[name](data)

    # conditional validator will take two arguments to evaluate

    if name == 'Conditional':
        result = validator.validate(value, value2)
    else:
        result = validator.validate(value)

    # If the cell value breaks the validation rule , append violations list
    if not result:
        violations.append(validator.getMessage())

    if len(violations) > 0:
        errors.append((coordinate, violations))

    # return result != False
    # result is the output of each validation for each cell
    # if not result:
    #     return False
    # else:
    #     return True
    return result


def set_settings(config):
    """function takes the config yaml file and converts it to dictionary
    """
    settings = {}

    # excludes are the columns that we won't validate

    excludes = []

    print("Get validation config " + config)
    try:
        stream = open(config, 'r')
    except IOError as e:
        print(e)
        exit(1)
    config = yaml.safe_load(stream)

    # Make sure that the yaml file follows the rules
    if 'validators' in config:
        if 'columns' in config.get('validators'):
            settings['validators'] = config.get('validators').get('columns')
        if 'header' in config.get('validators'):
            settings['header_validators'] = config.get('validators').get('header')
    else:
        return False

    if 'default' in config.get('validators'):
        settings['defaultValidator'] = config.get('validators').get('default')[0]
    else:
        settings['defaultValidator'] = None

    if 'excludes' in config:
        for column in config.get('excludes'):
            excludes.append(column_index_from_string(column))
        settings['excludes'] = excludes
    else:
        settings['excludes'] = []

    if 'range' in config:
        settings['range'] = config.get('range')[0] + "1:" + config.get('range')[1]
    else:
        settings['range'] = None

    if 'header' in config:
        settings['header'] = config.get('header')
    else:
        settings['header'] = None

    return settings


def mark_errors(errors, excelFile, sheetName, tmpDir, printErrors=False, noSizeLimit=False):
    """ Function takes the error lists (coordinates,violations) , excel file , sheet name
    output directory
    """
    print("")
    if printErrors is not None:
        for error in errors:
            print(f"Broken Excel cell: {error[0]} [{error[1]}]")
        return

    # Checking size of the file
    with yaspin(text="Checking Excel file size") as sp:
        file_too_big = os.path.getsize(excelFile) > 10485760
        if file_too_big is True and noSizeLimit is False:
            return -1
        sp.ok("✅ ")

    # open Excel file
    with yaspin(text="Setting up Excel file") as sp:
        error_file_name = "errors_" + time.strftime("%Y-%m-%d") + "_" + str(
            int(time.time())) + "_" + os.path.basename(
            excelFile)
        new_file = os.path.join(tmpDir, error_file_name)
        file_name, file_extension = os.path.splitext(excelFile)

        if file_extension == '.xlsm':
            try:
                wb = load_workbook(excelFile, keep_vba=True, data_only=True)
            except Exception as ex:
                sp.fail(f"Exception encountered: {ex}")
        else:
            try:
                wb = load_workbook(excelFile, data_only=True)
            except Exception as ex:
                sp.fail(f"Exception encountered: {ex}")

        creator = wb.properties.creator
        ws = wb[sheetName]

        # fill the error values with red pattern
        red_fill = PatternFill(start_color='FFFF0000',
                               end_color='FFFF0000',
                               fill_type='solid')
        sp.ok("✅ ")

    with alive_bar(len(errors), title="Processing errors", bar='classic', elapsed=False, stats=False) as bar:
        for error in errors:
            bar.text = f"Error found at: {error[0]}"
            if len(error[0]) == 2:
                cell = ws[error[0]]
                if printErrors:
                    cell.value = ','.join(error[1])
                cell.fill = red_fill
            else:
                message_split = error[0].split(" ")
                if message_split[0] == "Row":
                    for cell in ws[message_split[1]]:
                        if hasattr(cell, 'column') and cell.column in settings['excludes']:
                            continue
                        cell.fill = red_fill
            bar()

    with yaspin(text="Logging errors to file") as sp:
        wb.create_sheet("Log")
        sheet = wb["Log"]
        sheet['A1'] = "Location"
        sheet['B1'] = "Validation error"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        sheet.insert_rows(2, len(errors) + 1)

        # TODO: Split this into two columns
        for idx, item in enumerate(errors):
            # sp.text(f"Writing to row {idx + 2}")
            sheet.cell(column=1, row=idx + 2, value=str(item[0]))
            sheet.cell(column=2, row=idx + 2, value=str(item[1]))

        # save error log excel file
        wb.properties.creator = creator

        try:
            wb.save(new_file)
            sp.text = new_file
            sp.ok(f"✅ Saved file at: ")
        except Exception as ex:
            sp.fail(f"Error saving file: {ex}")
            exit(1)

    return new_file


def validate(settings, excelFile, sheetName, tmpDir, printErrors=False, noSizeLimit=False):
    """the main function of validations, takes settings dictionary (validations)
    and returns the validation result
    """
    print("Validate Excel Sheet " + sheetName)

    errors = []
    # open Excel file
    print("Parse Excel file")
    wb = load_workbook(excelFile, keep_vba=True, data_only=True, read_only=True)
    ws = wb[sheetName]

    progress_bar = Bar('Processing', max=ws.max_row)

    if 'range' in settings and settings['range'] is not None:
        settings['range'] = settings['range'] + str(ws.max_row)

    # range now equals A1:D(150) for example

    # iterate Excel sheet
    row_counter = 1
    header_row = 0
    if settings['header'] is not None:
        header_row = settings['header']
    for row in ws.iter_rows(settings['range']):
        progress_bar.next()
        column_counter = 0
        # do not parse empty rows
        if is_empty(row):
            continue
        if row_counter == header_row:
            coordinates = "Row " + str(header_row)
            # Get the validators for the headers from the yaml file
            for type in settings['header_validators']:
                name = list(type.keys())[0]
                if name == 'Order':
                    value = []
                    try:
                        for cell in ws[header_row]:
                            if hasattr(cell, 'column') and cell.column in settings['excludes']:
                                continue
                            current_row_coords = cell.coordinate
                            value.append(cell.value)
                    except ValueError:
                        errors.append((current_row_coords, ValueError))
                    res = is_valid(type, value, coordinates, errors)
                    if not res:
                        break

        for cell in row:
            column_counter = column_counter + 1
            try:
                value = cell.value
            except ValueError:
                # case when it is not possible to read value at all from any reason
                column = get_column_letter(column_counter)
                coordinates = "%s%d" % (column, row_counter)
                errors.append((coordinates, ValueError))

            # skip excludes column
            if hasattr(cell, 'column') and cell.column in settings['excludes']:
                continue

            # if cell row is header row, skip it (header row is already validated)
            if hasattr(cell, 'column') and cell.row == header_row:
                continue

            column = get_column_letter(column_counter)

            # TODO: Implement skip header row number
            # This will solve the mismatch in errors here vs the source repo
            # Replace with cell.coordinate?
            coordinates = "%s%d" % (column, row_counter)

            if column in settings['validators']:
                for type in settings['validators'][column]:
                    name = list(type.keys())[0]  # not-blank, Regex, Length
                    if name != 'Conditional':
                        res = is_valid(type, value, coordinates, errors)
                    else:
                        field_b = list(type.values())[0]['fieldB']
                        value2 = ws[field_b + str(row_counter)].value
                        res = is_valid(type, value, coordinates, errors, value2)
                    if not res:
                        break

            elif settings['defaultValidator'] is not None:
                is_valid(settings['defaultValidator'], value, coordinates, errors)
        row_counter += 1
    progress_bar.finish()

    print("Found %d error(s)" % len(errors))
    if len(errors) > 0:
        return mark_errors(errors, excelFile, sheetName, tmpDir, printErrors, noSizeLimit)

    return True


def is_empty(row):
    """ function to get if the row is empty or not
    """
    for cell in row:
        if cell.value:
            return False

    return True


if __name__ == '__main__':

    faulthandler.enable()

    parser = argparse.ArgumentParser(description='Mark validation errors in Excel sheet.')
    parser.add_argument('config', metavar='config', help='Path to YAML config file')
    parser.add_argument('file', metavar='file', help='Path to excel sheet file')
    parser.add_argument('sheetName', metavar='sheetName', help='Excel Sheet Name')
    parser.add_argument('tmpDir', metavar='tmpDir', help='Temporary directory path')
    parser.add_argument('--errors', metavar='errors',
                        help='Print errors messages without generating excel file with errors')
    parser.add_argument('--no-file-size-limit', metavar='size', help='Switch off file size limit. Use with care')
    args = parser.parse_args()

    settings = set_settings(args.config)

    excel_folder = os.path.dirname(args.file)

    if not os.path.exists(args.tmpDir):
        excel_folder = os.path.dirname(args.file)
        os.chdir(excel_folder)
        os.makedirs(os.path.join(os.getcwd(), "temp"), exist_ok=True)
        args.tmpDir = os.path.join(os.getcwd(), "temp")

    if not settings:
        sys.exit("Incorrect config file " + args.config)

    results = validate(settings, args.file, args.sheetName, args.tmpDir, args.errors, args.no_file_size_limit)

    # if result = True that means file is originally true and all values are correct
    # if result != True and not equal None, get result file name
    # if results == -1 File is too large , Exit

    if not results:
        if results and results != -1:
            sys.exit("Validation errors store in: [[" + results + "]]")
        elif results == -1:
            sys.exit("File is too big to generate annotated Excel file")

    sys.exit(0)
